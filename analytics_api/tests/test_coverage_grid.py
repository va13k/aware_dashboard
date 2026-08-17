"""The two coverage grids, as the heatmap asks for them.

The arithmetic underneath is covered in test_coverage_matrix.py and
test_sensor_rates.py. What these establish is the round trip the grid depends on
and nothing else checks: the rollup answers per *table* and a grid is read per
*sensor*, so a sensor stored across two tables has to arrive as one row of
numbers, and a table no sensor claims must not appear as a row at all.

Then the two states that make the grid honest rather than alarming. A device is
only expected to be sending inside an enrolment window, so a column before it
joined is neutral rather than a gap. And a required sensor that reported nothing
still gets a row on the device grid, because its absence is the finding.
"""

import csv
import io
import zipfile

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.routers import coverage as coverage_router
from app.services import coverage_matrix, sensor_rates, sensor_requirements

DEVICE = "phone-a"
QUIET = "phone-b"
#: 17 March 2026, UTC midnight. In the past, because an enrolment window left
#: open runs to *now* — a grid drawn over tomorrow would correctly hold nothing.
DAY_START = 1_773_705_600_000


def android_table(sensor: str) -> str:
    tables = coverage_router.ANDROID_EXPORT_MODELS
    return coverage_router.sensor_tables.tables_for(tables, sensor)[0]


@pytest.fixture
def grid(monkeypatch):
    """Lets a test say what the rollup holds and who was enrolled when.

    `holdings` is keyed per platform and then the way the rollup is —
    `(device, table)` — each entry a count per bucket, so a test states the shape
    the grid should read rather than the shape it wants back.
    """
    state = {
        "holdings": {"android": {}, "ios": {}},
        "windows": {DEVICE: [{"joined_at": 0, "left_at": None}]},
        "first_record": {},
        "required": ["battery", "bluetooth", "wifi"],
        "settings": {"frequency_bluetooth": 60, "frequency_wifi": 60},
    }

    async def bucketed_by_table(db, model, buckets, tables=None, device_id=None):
        platform = "ios" if "Ios" in model.__name__ else "android"
        selected = {}
        for (device, table), counts in state["holdings"][platform].items():
            if tables is not None and table not in tables:
                continue
            if device_id is not None and device != device_id:
                continue
            selected[(device, table)] = list(counts)[: len(buckets)]
        return selected

    async def stored_windows(db, model, device_id=None):
        if device_id is None:
            return {key: list(value) for key, value in state["windows"].items()}
        found = state["windows"].get(device_id)
        return {device_id: list(found)} if found else {}

    async def first_record_by_device(db, model):
        return dict(state["first_record"])

    def study_requirements():
        return {
            platform: sensor_requirements.PlatformRequirements(
                platform=platform,
                available=True,
                sensors=[
                    sensor_requirements.SensorRequirement(key, True)
                    for key in state["required"]
                ],
                required_sensor_count=len(state["required"]),
            )
            for platform in ("android", "ios")
        }

    def study_rates():
        return {
            platform: sensor_rates.rates_for(platform, state["settings"])
            for platform in ("android", "ios")
        }

    monkeypatch.setattr(
        coverage_router.coverage_matrix, "bucketed_by_table", bucketed_by_table
    )
    monkeypatch.setattr(coverage_router.enrolment, "stored_windows", stored_windows)
    monkeypatch.setattr(
        coverage_router.enrolment, "first_record_by_device", first_record_by_device
    )
    monkeypatch.setattr(
        coverage_router.sensor_requirements, "study_requirements", study_requirements
    )
    monkeypatch.setattr(coverage_router.sensor_rates, "study_rates", study_rates)
    return state


@pytest.fixture
def client(grid):
    app = FastAPI()
    app.include_router(coverage_router.router)

    async def session():
        yield None

    app.dependency_overrides[coverage_router.get_android_db] = session
    app.dependency_overrides[coverage_router.get_ios_db] = session
    with TestClient(app) as test_client:
        yield test_client
    app.dependency_overrides.clear()


def hours(client, params=""):
    return client.get(
        f"/coverage/study?level=hour&anchor={DAY_START}&platform=android&tz=UTC{params}"
    ).json()


def row_for(body, device):
    return next(row for row in body["rows"] if row["device_id"] == device)


def test_an_hour_grid_has_a_column_per_hour_of_the_anchors_day(client):
    body = hours(client)

    assert body["level"] == "hour"
    assert len(body["buckets"]) == 24
    assert body["buckets"][0]["from"] == DAY_START
    assert body["timezone"] == "UTC"


def test_a_level_names_what_it_drills_into(client):
    assert hours(client)["drills_into"] is None
    body = client.get(f"/coverage/study?level=month&anchor={DAY_START}").json()
    assert body["drills_into"] == "day"
    assert len(body["buckets"]) == 12


def test_an_unknown_level_is_refused(client):
    assert client.get("/coverage/study?level=fortnight").status_code == 422


def test_a_selected_sensor_reads_its_own_counts_per_bucket(client, grid):
    counts = [0] * 24
    counts[9] = 60
    grid["holdings"]["android"][(DEVICE, android_table("bluetooth"))] = counts

    body = hours(client, "&sensor=bluetooth")
    cells = row_for(body, DEVICE)["cells"]

    assert cells[9]["records"] == 60
    assert cells[8]["records"] == 0
    assert body["max_records"] == 60


def test_a_sensor_stored_in_two_tables_arrives_as_one_row(client, grid):
    """iOS `wifi` is why the rollup is keyed by table and a grid is not: its rows
    live in `sensor_wifi` and `wifi`, each with its own `_id` sequence, and a row
    of the grid means the sensor rather than one of its tables."""
    first = [0] * 24
    second = [0] * 24
    first[3] = 40
    second[3] = 20
    grid["holdings"]["ios"][(DEVICE, "sensor_wifi")] = first
    grid["holdings"]["ios"][(DEVICE, "wifi")] = second

    body = client.get(
        f"/coverage/study?level=hour&anchor={DAY_START}&platform=ios&sensor=wifi&tz=UTC"
    ).json()

    assert row_for(body, DEVICE)["cells"][3]["records"] == 60


def test_a_table_no_sensor_claims_is_not_counted(client, grid):
    """Bookkeeping tables share the rollup with sensor data and must not appear as
    a sensor's records under a guess."""
    counts = [0] * 24
    counts[5] = 99
    grid["holdings"]["android"][(DEVICE, "aware_log")] = counts

    body = hours(client)
    cells = row_for(body, DEVICE)["cells"]

    assert cells[5]["records"] == 0
    assert body["max_records"] == 0


def test_an_unknown_sensor_returns_no_rows(client):
    body = hours(client, "&sensor=telepathy")

    assert body["rows"] == []


def test_columns_before_a_device_joined_expect_nothing(client, grid):
    """The left edge of every grid: a study's first hours belong to whoever had
    joined by then, and read as neutral for everyone else."""
    grid["windows"][DEVICE] = [{"joined_at": DAY_START + 10 * 3_600_000, "left_at": None}]
    grid["holdings"]["android"][(DEVICE, android_table("bluetooth"))] = [0] * 24

    cells = row_for(hours(client, "&sensor=bluetooth"), DEVICE)["cells"]

    assert cells[0]["state"] == coverage_matrix.NOT_EXPECTED
    assert cells[9]["state"] == coverage_matrix.NOT_EXPECTED
    assert cells[10]["state"] == coverage_matrix.MISSING


def test_a_gap_between_enrolments_expects_nothing(client, grid):
    grid["windows"][DEVICE] = [
        {"joined_at": DAY_START, "left_at": DAY_START + 4 * 3_600_000},
        {"joined_at": DAY_START + 8 * 3_600_000, "left_at": None},
    ]
    grid["holdings"]["android"][(DEVICE, android_table("bluetooth"))] = [0] * 24

    cells = row_for(hours(client, "&sensor=bluetooth"), DEVICE)["cells"]

    assert cells[2]["state"] == coverage_matrix.MISSING
    assert cells[6]["state"] == coverage_matrix.NOT_EXPECTED
    assert cells[9]["state"] == coverage_matrix.MISSING


def test_a_device_with_a_window_and_no_data_still_gets_a_row(client, grid):
    """An empty row inside an enrolment is the clearest thing the view can show."""
    grid["windows"][QUIET] = [{"joined_at": 0, "left_at": None}]

    body = hours(client, "&sensor=bluetooth")

    assert {row["device_id"] for row in body["rows"]} == {DEVICE, QUIET}
    assert all(
        cell["state"] == coverage_matrix.MISSING for cell in row_for(body, QUIET)["cells"]
    )


def test_a_device_with_no_window_falls_back_to_its_first_record(client, grid):
    """iOS never uploads its study state, so without this every iPhone row would
    read as time nothing was expected."""
    grid["windows"] = {}
    grid["first_record"] = {DEVICE: DAY_START + 6 * 3_600_000}
    grid["holdings"]["android"][(DEVICE, android_table("bluetooth"))] = [0] * 24

    cells = row_for(hours(client, "&sensor=bluetooth"), DEVICE)["cells"]

    assert cells[5]["state"] == coverage_matrix.NOT_EXPECTED
    assert cells[6]["state"] == coverage_matrix.MISSING


def test_a_cell_carries_what_it_was_judged_against(client, grid):
    """A colour on its own does not say what it was compared with, and for a
    sensor whose configured and delivered rates differ by orders of magnitude
    that is the number worth seeing."""
    counts = [0] * 24
    counts[1] = 60
    grid["holdings"]["android"][(DEVICE, android_table("bluetooth"))] = counts

    cell = row_for(hours(client, "&sensor=bluetooth"), DEVICE)["cells"][1]

    assert cell["expected"] == 60
    assert cell["basis"] == sensor_rates.SCANNED
    assert cell["floor"] is True


def test_no_sensor_selected_counts_how_many_required_sensors_reported(client, grid):
    counts = [0] * 24
    counts[2] = 5
    grid["holdings"]["android"][(DEVICE, android_table("battery"))] = counts

    cell = row_for(hours(client), DEVICE)["cells"][2]

    assert (cell["reporting"], cell["required"]) == (1, 3)
    assert cell["fraction"] == round(1 / 3, 4)


def test_the_required_list_leaves_out_streams_with_no_table(client, grid):
    """A required setting with no table behind it cannot contribute to "how much
    of what we asked for arrived", and would hold every cell below full."""
    grid["required"] = ["battery", "esm-scheduler-that-does-not-exist"]

    body = hours(client)

    assert body["required_sensors"]["android"] == ["battery"]


def test_both_platforms_answer_one_grid(client, grid):
    grid["holdings"]["android"][(DEVICE, android_table("battery"))] = [1] * 24

    body = client.get(f"/coverage/study?level=hour&anchor={DAY_START}&tz=UTC").json()

    assert body["platforms"] == ["android", "ios"]
    assert row_for(body, DEVICE)["platform"] == "android"


def test_an_unknown_platform_is_refused(client):
    assert client.get("/coverage/study?platform=symbian").status_code == 404


def device_grid(client, params=""):
    return client.get(
        f"/coverage/device/android/{DEVICE}?level=hour&anchor={DAY_START}&tz=UTC{params}"
    ).json()


def test_the_device_grid_has_a_row_per_sensor(client, grid):
    counts = [0] * 24
    counts[7] = 12
    grid["holdings"]["android"][(DEVICE, android_table("battery"))] = counts

    body = device_grid(client)
    battery = next(row for row in body["rows"] if row["sensor"] == "battery")

    assert body["device_id"] == DEVICE
    assert battery["cells"][7]["records"] == 12
    assert battery["required"] is True


def test_a_required_sensor_that_reported_nothing_still_gets_a_row(client, grid):
    """Its absence is the finding, not a reason to leave it off the grid."""
    body = device_grid(client)

    assert {row["sensor"] for row in body["rows"]} >= {"battery", "bluetooth", "wifi"}
    quiet = next(row for row in body["rows"] if row["sensor"] == "wifi")
    assert quiet["records"] == 0


def test_a_sensor_reporting_outside_the_config_is_shown_unrequired(client, grid):
    """A sensor switched off in the config but still uploading is worth seeing."""
    grid["required"] = ["battery"]
    grid["holdings"]["android"][(DEVICE, android_table("bluetooth"))] = [1] * 24

    body = device_grid(client)
    extra = next(row for row in body["rows"] if row["sensor"] == "bluetooth")

    assert extra["required"] is False
    assert extra["records"] == 24


def test_the_device_grid_carries_the_rate_each_row_is_judged_against(client, grid):
    body = device_grid(client)
    scanned = next(row for row in body["rows"] if row["sensor"] == "bluetooth")
    event = next(row for row in body["rows"] if row["sensor"] == "battery")

    assert scanned["expected_per_hour"] == 60
    assert scanned["basis"] == sensor_rates.SCANNED
    assert event["expected_per_hour"] is None
    assert event["basis"] == sensor_rates.EVENT


def test_the_device_grid_reads_only_its_own_device(client, grid):
    grid["windows"][QUIET] = [{"joined_at": 0, "left_at": None}]
    grid["holdings"]["android"][(QUIET, android_table("battery"))] = [9] * 24

    body = device_grid(client)
    battery = next(row for row in body["rows"] if row["sensor"] == "battery")

    assert battery["records"] == 0


def test_an_unknown_platform_has_no_device_grid(client):
    assert client.get(f"/coverage/device/symbian/{DEVICE}").status_code == 404


def matrix(client, params=""):
    """The matrix archive, as `{member: [rows]}`."""
    response = client.get(
        f"/coverage/matrix?from_ts={DAY_START}&to_ts={DAY_START + 3 * 3_600_000}"
        f"&platform=android&tz=UTC{params}"
    )
    assert response.status_code == 200, response.text
    bundle = zipfile.ZipFile(io.BytesIO(response.content))
    return {
        name: list(csv.reader(io.StringIO(bundle.read(name).decode("utf-8"))))
        for name in bundle.namelist()
    }


def test_the_matrix_has_a_file_per_sensor_with_devices_down_and_hours_across(
    client, grid
):
    """The reference spreadsheet's layout, which is the point of producing it: a
    sheet per sensor, a device per row, an hour per column."""
    grid["holdings"]["android"][(DEVICE, android_table("battery"))] = [0, 4, 0, 9]

    members = matrix(client)
    header, *rows = members["android/battery.csv"]

    assert header[0] == "device_id"
    assert header[1] == "start"
    assert header[-1] == "covered_hours"
    # Four hour columns between the two bounds, inclusive of the hour each lands in.
    assert len(header) == 4 + 3
    assert rows[0][0] == DEVICE


def test_a_covered_hour_is_marked_and_an_empty_one_left_blank(client, grid):
    """1-or-blank, so the file can be diffed against the reference directly."""
    grid["holdings"]["android"][(DEVICE, android_table("battery"))] = [0, 4, 0, 9]

    (row,) = matrix(client)["android/battery.csv"][1:]

    assert row[2:6] == ["", "1", "", "1"]
    assert row[-1] == "2"


def test_the_start_column_names_the_first_covered_hour(client, grid):
    grid["holdings"]["android"][(DEVICE, android_table("battery"))] = [0, 4, 0, 0]

    (row,) = matrix(client)["android/battery.csv"][1:]

    assert row[1].endswith("01:00")


def test_counts_keep_the_magnitude_the_reference_layout_drops(client, grid):
    grid["holdings"]["android"][(DEVICE, android_table("battery"))] = [0, 4, 0, 9]

    (row,) = matrix(client, "&values=counts")["android/battery.csv"][1:]

    assert row[2:6] == ["", "4", "", "9"]


def test_a_sensor_no_phone_reported_gets_no_file(client, grid):
    """An archive of empty sheets is worse than an archive of what exists."""
    grid["holdings"]["android"][(DEVICE, android_table("battery"))] = [1, 1, 1, 1]

    members = matrix(client)

    assert "android/battery.csv" in members
    assert "android/bluetooth.csv" not in members


def test_the_archive_explains_itself(client, grid):
    """It leaves the dashboard and sits in a downloads folder, where the window
    and the timezone the columns were cut in are no longer on screen."""
    grid["holdings"]["android"][(DEVICE, android_table("battery"))] = [1, 0, 0, 0]

    readme = matrix(client)["README.txt"]
    text = "\n".join(",".join(row) for row in readme)

    assert "Timezone: UTC" in text
    assert "Hour columns: 4" in text


def test_a_window_without_both_bounds_is_refused(client):
    assert client.get(f"/coverage/matrix?from_ts={DAY_START}").status_code == 422


def test_an_absurd_window_is_refused_rather_than_produced(client):
    """A year of hour columns is a mistyped request, not a large study."""
    response = client.get(
        f"/coverage/matrix?from_ts={DAY_START}&to_ts={DAY_START + 400 * 86_400_000}"
    )

    assert response.status_code == 422
    assert "at most" in response.json()["detail"]


def test_an_unknown_value_mode_is_refused(client):
    response = client.get(
        f"/coverage/matrix?from_ts={DAY_START}&to_ts={DAY_START + 3_600_000}&values=vibes"
    )

    assert response.status_code == 422


def workbook(client, path):
    response = client.get(path)
    assert response.status_code == 200, response.text
    assert "spreadsheetml" in response.headers["content-type"]
    return load_workbook(io.BytesIO(response.content))


def test_the_study_workbook_holds_the_grid_that_was_on_screen(client, grid):
    """Same parameters as the grid endpoint, so the file is the view rather than a
    layout of its own."""
    counts = [0] * 24
    counts[9] = 60
    grid["holdings"]["android"][(DEVICE, android_table("bluetooth"))] = counts

    book = workbook(
        client,
        f"/coverage/study.xlsx?level=hour&anchor={DAY_START}&platform=android"
        "&sensor=bluetooth&tz=UTC",
    )
    sheet = book["Coverage"]

    assert sheet["A1"].value == "Device"
    # 24 hour columns, a label column and a total column.
    assert sheet.max_column == 26
    assert sheet.cell(row=2, column=1).value.startswith(DEVICE)
    assert sheet.cell(row=2, column=11).value == 60


def test_the_study_workbook_follows_the_level(client, grid):
    book = workbook(
        client, f"/coverage/study.xlsx?level=month&anchor={DAY_START}&platform=android"
    )

    # Twelve months, plus the label and total columns.
    assert book["Coverage"].max_column == 14
    assert book["Coverage"].cell(row=1, column=2).value == "Jan"


def test_the_study_workbook_records_the_view_it_came_from(client, grid):
    book = workbook(
        client,
        f"/coverage/study.xlsx?level=day&anchor={DAY_START}&platform=android&tz=UTC",
    )
    key = book["Key"]
    written = {
        key.cell(row=line, column=2).value: key.cell(row=line, column=3).value
        for line in range(1, key.max_row + 1)
    }

    assert written["Level"] == "day buckets"
    assert written["Timezone"] == "UTC"
    assert written["Platforms"] == "android"


def test_the_device_workbook_holds_a_sensor_per_row(client, grid):
    counts = [0] * 24
    counts[7] = 12
    grid["holdings"]["android"][(DEVICE, android_table("battery"))] = counts

    book = workbook(
        client,
        f"/coverage/device/android/{DEVICE}/workbook.xlsx?level=hour&anchor={DAY_START}&tz=UTC",
    )
    sheet = book["Coverage"]
    labels = [sheet.cell(row=line, column=1).value for line in range(2, sheet.max_row)]

    assert sheet["A1"].value == "Sensor"
    assert "battery" in labels
    battery = labels.index("battery") + 2
    assert sheet.cell(row=battery, column=9).value == 12


def test_the_device_workbook_marks_a_sensor_the_config_did_not_ask_for(client, grid):
    grid["required"] = ["battery"]
    grid["holdings"]["android"][(DEVICE, android_table("bluetooth"))] = [1] * 24

    book = workbook(
        client,
        f"/coverage/device/android/{DEVICE}/workbook.xlsx?level=hour&anchor={DAY_START}&tz=UTC",
    )
    sheet = book["Coverage"]
    labels = [sheet.cell(row=line, column=1).value for line in range(2, sheet.max_row)]

    assert "bluetooth (extra)" in labels


def test_an_unknown_platform_has_no_device_workbook(client):
    assert client.get(f"/coverage/device/symbian/{DEVICE}/workbook.xlsx").status_code == 404


def test_an_unknown_level_is_refused_by_the_workbook(client):
    assert client.get("/coverage/study.xlsx?level=fortnight").status_code == 422


def test_an_empty_matrix_archive_says_so(client):
    """The window a grid opens on may hold nothing, and an archive of one note is
    otherwise indistinguishable from a broken download."""
    body = matrix(client)["README.txt"]
    text = "\n".join(",".join(row) for row in body)

    assert "Sensor files: 0" in text
    assert "No sensor reported anything inside this window" in text
