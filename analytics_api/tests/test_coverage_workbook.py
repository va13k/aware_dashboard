"""The coverage grid as a spreadsheet.

The file leaves the dashboard and gets circulated, so what it claims has to hold
up away from the interface that produced it. Three things are checked.

Each cell carries the count *and* the fill its band gives it, because the colour is
half of what a coverage grid says — 484 records means one thing against an
expectation of 180,000 and another against 500.

A bucket outside enrolment is left empty rather than zeroed. A zero is a
measurement and would be averaged as one by whatever the researcher does next; an
empty cell says nothing was asked of that period.

Totals are formulas rather than figures, so a reader who sorts, filters or deletes
a row still has a spreadsheet that adds up.
"""

import io

from openpyxl import load_workbook

from app.services import coverage_matrix, coverage_workbook

BUCKETS = [
    {"key": "h0", "label": "00", "from": 0, "to": 1},
    {"key": "h1", "label": "01", "from": 1, "to": 2},
    {"key": "h2", "label": "02", "from": 2, "to": 3},
]


def cell(band, records, **extra):
    return {"band": band, "records": records, "hours": 1, **extra}


def sheets(rows, row_header="Device", about=None):
    content = coverage_workbook.build(
        buckets=BUCKETS,
        rows=rows,
        row_header=row_header,
        about=about or [("Study", "Test study")],
    )
    book = load_workbook(io.BytesIO(content))
    return book


def test_the_grid_holds_a_row_per_row_and_a_column_per_bucket():
    book = sheets(
        [
            {
                "label": "phone-a",
                "cells": [
                    cell(coverage_matrix.BAND_EXPECTED, 60, expected=60),
                    cell(coverage_matrix.BAND_SHORT, 5, expected=60),
                    cell(coverage_matrix.BAND_NONE, 0, expected=60),
                ],
            }
        ]
    )
    grid = book["Coverage"]

    assert grid["A1"].value == "Device"
    assert [grid.cell(row=1, column=index + 2).value for index in range(3)] == [
        "00",
        "01",
        "02",
    ]
    assert grid["A2"].value == "phone-a"
    assert [grid.cell(row=2, column=index + 2).value for index in range(3)] == [60, 5, 0]


def test_each_band_paints_its_own_fill():
    """A red cell in the spreadsheet is red for the reason it is red on screen."""
    book = sheets(
        [
            {
                "label": "phone-a",
                "cells": [
                    cell(coverage_matrix.BAND_SHORT, 1),
                    cell(coverage_matrix.BAND_EXPECTED, 1),
                    cell(coverage_matrix.BAND_OVER, 1),
                ],
            }
        ]
    )
    grid = book["Coverage"]
    fills = [grid.cell(row=2, column=index + 2).fill.fgColor.rgb for index in range(3)]

    assert len(set(fills)) == 3
    for band, painted in zip(
        (
            coverage_matrix.BAND_SHORT,
            coverage_matrix.BAND_EXPECTED,
            coverage_matrix.BAND_OVER,
        ),
        fills,
    ):
        assert coverage_workbook.BAND_FILL[band] in painted


def test_a_bucket_outside_enrolment_is_left_empty_rather_than_zeroed():
    """A zero is a measurement; nothing was measured here."""
    book = sheets(
        [
            {
                "label": "phone-a",
                "cells": [
                    cell(coverage_matrix.BAND_BLANK, 0),
                    cell(coverage_matrix.BAND_NONE, 0),
                    cell(coverage_matrix.BAND_EXPECTED, 9),
                ],
            }
        ]
    )
    grid = book["Coverage"]

    assert grid.cell(row=2, column=2).value is None
    assert grid.cell(row=2, column=3).value == 0
    assert grid.cell(row=2, column=4).value == 9


def test_a_cell_carries_what_it_was_judged_against():
    """The count alone cannot say whether it is healthy, and a spreadsheet has
    nowhere else to put the expectation."""
    book = sheets(
        [
            {
                "label": "phone-a",
                "cells": [
                    cell(coverage_matrix.BAND_SHORT, 484, expected=180000),
                    cell(coverage_matrix.BAND_OVER, 348, expected=13, floor=True),
                    cell(coverage_matrix.BAND_UNJUDGED, 7),
                ],
            }
        ]
    )
    grid = book["Coverage"]

    assert "180,000" in grid.cell(row=2, column=2).comment.text
    assert "at least" in grid.cell(row=2, column=3).comment.text
    assert "bounds the scans" in grid.cell(row=2, column=3).comment.text
    assert "no configured rate" in grid.cell(row=2, column=4).comment.text


def test_a_gated_cell_says_its_figure_is_a_ceiling():
    """The band's own wording denies a rate the researcher can see in the config,
    which would send them after the missing setting instead of the filter."""
    book = sheets(
        [
            {
                "label": "phone-a",
                "cells": [
                    cell(
                        coverage_matrix.BAND_UNJUDGED,
                        200,
                        expected=180000,
                        ceiling=True,
                    ),
                    cell(coverage_matrix.BAND_NONE, 0),
                    cell(coverage_matrix.BAND_NONE, 0),
                ],
            }
        ]
    )
    note = book["Coverage"].cell(row=2, column=2).comment.text

    assert "at most 180,000" in note
    assert "filters this sensor" in note
    assert "no configured rate" not in note


def test_a_partly_covered_bucket_says_how_long_it_was_enrolled():
    book = sheets(
        [
            {
                "label": "phone-a",
                "cells": [
                    {
                        "band": coverage_matrix.BAND_EXPECTED,
                        "records": 5,
                        "hours": 0.25,
                        "expected": 15,
                    },
                    cell(coverage_matrix.BAND_NONE, 0),
                    cell(coverage_matrix.BAND_NONE, 0),
                ],
            }
        ]
    )

    assert "15 min" in book["Coverage"].cell(row=2, column=2).comment.text


def test_a_row_total_sums_its_own_buckets():
    book = sheets(
        [
            {
                "label": "phone-a",
                "cells": [cell(coverage_matrix.BAND_EXPECTED, n) for n in (1, 2, 3)],
            }
        ]
    )
    grid = book["Coverage"]

    assert grid.cell(row=1, column=5).value == "Total"
    assert grid.cell(row=2, column=5).value == "=SUM(B2:D2)"


def test_a_column_total_sums_every_row():
    book = sheets(
        [
            {"label": "a", "cells": [cell(coverage_matrix.BAND_EXPECTED, 1)] * 3},
            {"label": "b", "cells": [cell(coverage_matrix.BAND_EXPECTED, 2)] * 3},
        ]
    )
    grid = book["Coverage"]

    assert grid.cell(row=4, column=1).value == "Total"
    assert grid.cell(row=4, column=2).value == "=SUM(B2:B3)"
    # The corner totals the row totals, so the grand total is there too.
    assert grid.cell(row=4, column=5).value == "=SUM(E2:E3)"


def test_the_totals_are_formulas_so_a_sorted_sheet_still_adds_up():
    book = sheets(
        [{"label": "a", "cells": [cell(coverage_matrix.BAND_EXPECTED, 5)] * 3}]
    )
    grid = book["Coverage"]

    assert str(grid.cell(row=3, column=2).value).startswith("=SUM(")
    assert str(grid.cell(row=2, column=5).value).startswith("=SUM(")


def test_the_key_sheet_explains_every_fill_the_grid_can_use():
    """The file gets read away from the legend the interface draws."""
    book = sheets([{"label": "a", "cells": [cell(coverage_matrix.BAND_NONE, 0)] * 3}])
    key = book["Key"]

    named = {
        key.cell(row=line, column=2).value
        for line in range(1, key.max_row + 1)
    }
    assert set(coverage_workbook.BAND_FILL) <= named


def test_the_key_sheet_records_which_view_the_grid_came_from():
    book = sheets(
        [{"label": "a", "cells": [cell(coverage_matrix.BAND_NONE, 0)] * 3}],
        about=[("Level", "day buckets"), ("Timezone", "Europe/Zurich")],
    )
    key = book["Key"]
    written = {
        key.cell(row=line, column=2).value: key.cell(row=line, column=3).value
        for line in range(1, key.max_row + 1)
    }

    assert written["Level"] == "day buckets"
    assert written["Timezone"] == "Europe/Zurich"


def test_the_header_row_and_labels_stay_put_when_scrolled():
    book = sheets([{"label": "a", "cells": [cell(coverage_matrix.BAND_NONE, 0)] * 3}])

    assert book["Coverage"].freeze_panes == "B2"


def test_a_grid_with_no_rows_still_produces_a_readable_sheet():
    """An empty period is a normal thing to ask about, and a file that fails to
    open is a worse answer than one that says there is nothing in it."""
    book = sheets([])
    grid = book["Coverage"]

    assert grid["A1"].value == "Device"
    assert grid.cell(row=2, column=1).value == "Total"
