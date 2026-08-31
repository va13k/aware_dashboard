"""The coverage grid as a spreadsheet: the same cells, colours and totals.

A researcher reading the grid on screen ends up wanting it in a file — to put in
a supervision meeting, to keep beside a dataset, to sort and annotate. What makes
this worth building rather than pointing at a CSV export is that the *colour* is
half of what the grid says: a count of 484 means one thing against an expectation
of 180,000 and another against 500, and a plain table of numbers drops the
comparison the view exists to make.

So the workbook carries three things per cell — the count, the fill its band gives
it, and a comment holding what it was judged against — plus a total down every
row and across every column.

Bands come from `coverage_matrix.band_for`, the same function the API serves to
the browser, so a red cell in the spreadsheet is red for the reason it is red on
screen. The fills here are the only thing this module decides, and they are the
Excel equivalents of the interface's own palette.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services import coverage_matrix as matrix

#: Band fills, matching the palette the interface draws with. Desaturated so a
#: sheet of them stays readable, and separated in lightness as well as hue so the
#: order survives a monochrome print.
BAND_FILL = {
    matrix.BAND_SHORT: "F4C7C3",
    matrix.BAND_MODERATE: "FBE2B7",
    matrix.BAND_EXPECTED: "C6E7CE",
    matrix.BAND_OVER: "C9DCF3",
    matrix.BAND_UNJUDGED: "DFE3E1",
    #: Expected, and nothing arrived. Left white, with the grid's own border.
    matrix.BAND_NONE: None,
    #: Nothing was expected. A wash, the way the interface draws it.
    matrix.BAND_BLANK: "F2F1EE",
}

#: `unjudged` where a rate does exist but the phone filters what reaches the
#: table. The band's own wording denies a rate that is right there in the config.
GATED_MEANING = "Arrived — the phone filters this sensor, so the amount is not judged"

BAND_MEANING = {
    matrix.BAND_SHORT: "Well under the rate the study config asks for",
    matrix.BAND_MODERATE: "Approaching the rate the study config asks for",
    matrix.BAND_EXPECTED: "At the rate the study config asks for",
    matrix.BAND_OVER: "Far above the rate the study config asks for",
    matrix.BAND_UNJUDGED: "Arrived — no configured rate to compare it with",
    matrix.BAND_NONE: "Expected, and nothing arrived",
    matrix.BAND_BLANK: "Nothing expected — outside this device's enrolment",
}

HEADER_FILL = PatternFill("solid", fgColor="EFE7DA")
TOTAL_FILL = PatternFill("solid", fgColor="E7E2D6")
HAIRLINE = Side(style="thin", color="D8D4CA")
CELL_BORDER = Border(left=HAIRLINE, right=HAIRLINE, top=HAIRLINE, bottom=HAIRLINE)

#: Wide enough for a bucket label and a six-figure count.
BUCKET_WIDTH = 11
LABEL_WIDTH = 42
TOTAL_WIDTH = 14


def _fill_for(band: str) -> PatternFill | None:
    colour = BAND_FILL.get(band)
    return PatternFill("solid", fgColor=colour) if colour else None


def _cell_note(cell: dict) -> str | None:
    """What the count was judged against, for the cell's comment.

    The number alone cannot say whether it is healthy, and a spreadsheet has
    nowhere else to put the expectation without adding a column per bucket.
    """
    band = cell.get("band")
    gated = band == matrix.BAND_UNJUDGED and cell.get("ceiling")
    lines = [GATED_MEANING if gated else BAND_MEANING.get(band, "")]

    if cell.get("required") is not None:
        lines.append(
            f"{cell.get('reporting', 0)} of {cell.get('required', 0)} "
            "required sensors reported"
        )
    elif cell.get("expected") is not None:
        verb = "about"
        if cell.get("floor"):
            verb = "at least"
        elif cell.get("ceiling"):
            verb = "at most"
        lines.append(f"Config implies {verb} {round(cell['expected']):,}")
        if cell.get("floor"):
            lines.append("That figure bounds the scans, not the rows each yields.")

    hours = cell.get("hours") or 0
    if 0 < hours < 1:
        lines.append(f"Enrolled for {round(hours * 60)} min of this bucket.")

    text = "\n".join(line for line in lines if line)
    return text or None


def _write_grid(sheet, buckets: list[dict], rows: list[dict], row_header: str) -> None:
    """The matrix itself: buckets across, rows down, totals on both edges."""
    sheet.cell(row=1, column=1, value=row_header).font = Font(bold=True)
    sheet.cell(row=1, column=1).fill = HEADER_FILL

    for index, bucket in enumerate(buckets):
        heading = sheet.cell(row=1, column=index + 2, value=bucket["label"])
        heading.font = Font(bold=True)
        heading.fill = HEADER_FILL
        heading.alignment = Alignment(horizontal="center")

    total_column = len(buckets) + 2
    heading = sheet.cell(row=1, column=total_column, value="Total")
    heading.font = Font(bold=True)
    heading.fill = TOTAL_FILL

    for offset, row in enumerate(rows):
        line = offset + 2
        label = sheet.cell(row=line, column=1, value=row["label"])
        label.alignment = Alignment(horizontal="left")

        for index, cell in enumerate(row["cells"]):
            written = sheet.cell(row=line, column=index + 2)
            # A bucket outside enrolment is left empty rather than zeroed: it is
            # not a measurement, and a zero would be counted as one by anything
            # averaging the column afterwards.
            if cell.get("band") != matrix.BAND_BLANK:
                written.value = cell.get("records", 0)
            written.number_format = "#,##0"
            written.alignment = Alignment(horizontal="center")
            written.border = CELL_BORDER
            fill = _fill_for(cell.get("band", ""))
            if fill is not None:
                written.fill = fill
            note = _cell_note(cell)
            if note:
                written.comment = Comment(note, "AWARE Dashboard")

        first = get_column_letter(2)
        last = get_column_letter(len(buckets) + 1)
        total = sheet.cell(
            row=line,
            column=total_column,
            value=f"=SUM({first}{line}:{last}{line})",
        )
        total.font = Font(bold=True)
        total.fill = TOTAL_FILL
        total.number_format = "#,##0"

    # Totals across the bottom, as formulas rather than figures, so a reader who
    # sorts or deletes a row gets a spreadsheet that still adds up.
    footer = len(rows) + 2
    label = sheet.cell(row=footer, column=1, value="Total")
    label.font = Font(bold=True)
    label.fill = TOTAL_FILL

    for index in range(len(buckets) + 1):
        letter = get_column_letter(index + 2)
        total = sheet.cell(
            row=footer,
            column=index + 2,
            value=f"=SUM({letter}2:{letter}{footer - 1})",
        )
        total.font = Font(bold=True)
        total.fill = TOTAL_FILL
        total.number_format = "#,##0"

    sheet.column_dimensions["A"].width = LABEL_WIDTH
    for index in range(len(buckets)):
        sheet.column_dimensions[get_column_letter(index + 2)].width = BUCKET_WIDTH
    sheet.column_dimensions[get_column_letter(total_column)].width = TOTAL_WIDTH
    # The header row and the labels stay put while a wide grid is scrolled.
    sheet.freeze_panes = "B2"


def _write_key(sheet, about: list[tuple[str, object]]) -> None:
    """What each fill means, and which view the sheet was taken from."""
    sheet.cell(row=1, column=1, value="What the colours mean").font = Font(bold=True)

    line = 2
    for band, meaning in BAND_MEANING.items():
        swatch = sheet.cell(row=line, column=1)
        fill = _fill_for(band)
        if fill is not None:
            swatch.fill = fill
        swatch.border = CELL_BORDER
        sheet.cell(row=line, column=2, value=band)
        sheet.cell(row=line, column=3, value=meaning)
        line += 1

    line += 1
    sheet.cell(row=line, column=1, value="This sheet shows").font = Font(bold=True)
    line += 1
    for name, value in about:
        sheet.cell(row=line, column=2, value=name)
        sheet.cell(row=line, column=3, value=value)
        line += 1

    sheet.column_dimensions["A"].width = 5
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 62


def build(
    buckets: list[dict],
    rows: list[dict],
    row_header: str,
    about: list[tuple[str, object]],
    sheet_title: str = "Coverage",
) -> bytes:
    """The workbook, as the bytes of an `.xlsx` file.

    Held in memory rather than streamed: the size follows the grid on screen, so a
    sheet is one row per device or per sensor however many millions of records
    those buckets hold.
    """
    book = Workbook()
    grid = book.active
    grid.title = sheet_title
    _write_grid(grid, buckets, rows, row_header)
    _write_key(book.create_sheet("Key"), about)

    stream = BytesIO()
    book.save(stream)
    return stream.getvalue()
